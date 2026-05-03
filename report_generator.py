"""Excel report generator for any strategy's backtest or live trading results.

Generates a professional multi-sheet Excel workbook from a BacktestResult,
including summary metrics, trade log, equity curve, P&L distribution,
monthly breakdown, and daily P&L sheets with embedded charts.

Requires: openpyxl
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from trading_framework.models import BacktestResult, Trade

IST = ZoneInfo("Asia/Kolkata")

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)


def _fmt_ts(ts: datetime | None) -> str:
    if ts is None:
        return ""
    return ts.astimezone(IST).strftime("%Y-%m-%d %H:%M")


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max_len + 3, 40)
        ws.column_dimensions[col_letter].width = adjusted


class ReportGenerator:
    """Generate a comprehensive Excel report from a BacktestResult."""

    def __init__(self, result: BacktestResult) -> None:
        self.result = result
        self.wb = Workbook()

    def generate(self, filepath: str = "backtest_report.xlsx") -> str:
        """Generate the complete Excel report and save to *filepath*."""
        self._create_summary_sheet()
        self._create_trade_log_sheet()
        self._create_equity_curve_sheet()
        self._create_pnl_distribution_sheet()
        self._create_monthly_breakdown_sheet()
        self._create_daily_pnl_sheet()

        # Iron Condor-specific sheets (only if strategy is iron_condor)
        if self.result.strategy_name == "iron_condor":
            self._create_iron_condor_metrics_sheet()
            self._create_expiry_cycle_sheet()

        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        self.wb.save(filepath)
        print(f"Report saved to {filepath}")
        return filepath

    # ------------------------------------------------------------------
    # Sheet 1 – Summary
    # ------------------------------------------------------------------

    def _create_summary_sheet(self) -> None:
        ws = self.wb.create_sheet("Summary")
        r = self.result

        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        strategy_title = r.strategy_name or "Strategy"
        title_cell.value = f"{strategy_title} – Backtest Summary"
        title_cell.font = TITLE_FONT

        row = 3

        def _metric(label: str, value, fmt: str | None = None) -> None:
            nonlocal row
            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            cell = ws.cell(row=row, column=2, value=value)
            if fmt:
                cell.number_format = fmt
            if isinstance(value, (int, float)):
                if value > 0:
                    cell.fill = GREEN_FILL
                elif value < 0:
                    cell.fill = RED_FILL
            row += 1

        _metric("Period", f"{r.start_date} – {r.end_date}" if r.start_date else "N/A")
        _metric("Total Trading Days", r.total_trading_days)
        row += 1

        _metric("Total Trades", r.total_trades)
        _metric("Winning Trades", r.winning_trades)
        _metric("Losing Trades", r.losing_trades)
        _metric("Breakeven Trades", r.breakeven_trades)
        _metric("Win Rate %", r.win_rate, "0.00")
        row += 1

        _metric("Total P&L (points)", r.total_pnl_points, "0.00")
        _metric("Total P&L (₹)", r.total_pnl_rupees, "#,##0.00")
        _metric("Avg Win (points)", r.avg_win_points, "0.00")
        _metric("Avg Loss (points)", r.avg_loss_points, "0.00")
        _metric("Max Win (points)", r.max_win_points, "0.00")
        _metric("Max Loss (points)", r.max_loss_points, "0.00")
        row += 1

        pf_display = r.profit_factor if r.profit_factor != float("inf") else "∞"
        _metric("Profit Factor", pf_display, "0.00" if isinstance(pf_display, float) else None)
        _metric("Risk/Reward Ratio", r.risk_reward_ratio, "0.00")
        _metric("Max Drawdown (points)", r.max_drawdown_points, "0.00")
        _metric("Max Drawdown (₹)", r.max_drawdown_rupees, "#,##0.00")
        row += 1

        signal_exits = sum(1 for t in r.trades if t.exit_reason == "signal")
        eod_exits = sum(1 for t in r.trades if t.exit_reason == "eod")
        data_exits = sum(1 for t in r.trades if t.exit_reason == "end_of_data")
        _metric("Exits – Signal", signal_exits)
        _metric("Exits – EOD", eod_exits)
        _metric("Exits – End of Data", data_exits)

        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 2 – Trade Log
    # ------------------------------------------------------------------

    def _create_trade_log_sheet(self) -> None:
        ws = self.wb.create_sheet("Trade Log")
        headers = [
            "Trade #", "Direction", "Entry Time", "Entry Price",
            "Exit Time", "Exit Price", "P&L Points", "P&L ₹",
            "Exit Reason", "Cumulative P&L Points",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        cumulative = 0.0
        for i, trade in enumerate(self.result.trades, 1):
            cumulative += trade.pnl_points
            row = i + 1
            ws.cell(row=row, column=1, value=trade.trade_id or i)
            ws.cell(row=row, column=2, value=trade.direction)
            ws.cell(row=row, column=3, value=_fmt_ts(trade.entry_time))
            ws.cell(row=row, column=4, value=round(trade.entry_price, 2)).number_format = "0.00"
            ws.cell(row=row, column=5, value=_fmt_ts(trade.exit_time))
            ws.cell(row=row, column=6, value=round(trade.exit_price, 2)).number_format = "0.00"

            pnl_cell = ws.cell(row=row, column=7, value=round(trade.pnl_points, 2))
            pnl_cell.number_format = "0.00"
            pnl_cell.fill = GREEN_FILL if trade.pnl_points > 0 else RED_FILL if trade.pnl_points < 0 else PatternFill()

            rupee_cell = ws.cell(row=row, column=8, value=round(trade.pnl_rupees, 2))
            rupee_cell.number_format = "#,##0.00"
            rupee_cell.fill = GREEN_FILL if trade.pnl_rupees > 0 else RED_FILL if trade.pnl_rupees < 0 else PatternFill()

            ws.cell(row=row, column=9, value=trade.exit_reason)
            ws.cell(row=row, column=10, value=round(cumulative, 2)).number_format = "0.00"

        if self.result.trades:
            last_col = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{last_col}{len(self.result.trades) + 1}"

        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 3 – Equity Curve
    # ------------------------------------------------------------------

    def _create_equity_curve_sheet(self) -> None:
        ws = self.wb.create_sheet("Equity Curve")

        ws.cell(row=1, column=1, value="Trade #").font = BOLD_FONT
        ws.cell(row=1, column=2, value="Cumulative P&L (points)").font = BOLD_FONT

        cumulative = 0.0
        for i, trade in enumerate(self.result.trades, 1):
            cumulative += trade.pnl_points
            ws.cell(row=i + 1, column=1, value=i)
            ws.cell(row=i + 1, column=2, value=round(cumulative, 2))

        if not self.result.trades:
            ws.cell(row=2, column=1, value="No trades")
            return

        chart = LineChart()
        chart.title = "Equity Curve"
        chart.x_axis.title = "Trade #"
        chart.y_axis.title = "Cumulative P&L (points)"
        chart.width = 20
        chart.height = 12
        chart.style = 10

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(self.result.trades) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(self.result.trades) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.line.width = 20000

        ws.add_chart(chart, "D2")
        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 4 – P&L Distribution
    # ------------------------------------------------------------------

    def _create_pnl_distribution_sheet(self) -> None:
        ws = self.wb.create_sheet("P&L Distribution")

        ws.cell(row=1, column=1, value="Trade #").font = BOLD_FONT
        ws.cell(row=1, column=2, value="P&L (points)").font = BOLD_FONT

        for i, trade in enumerate(self.result.trades, 1):
            ws.cell(row=i + 1, column=1, value=i)
            ws.cell(row=i + 1, column=2, value=round(trade.pnl_points, 2))

        if not self.result.trades:
            ws.cell(row=2, column=1, value="No trades")
            return

        chart = BarChart()
        chart.title = "Trade P&L Distribution"
        chart.x_axis.title = "Trade #"
        chart.y_axis.title = "P&L (points)"
        chart.width = 20
        chart.height = 12
        chart.style = 10

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(self.result.trades) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(self.result.trades) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        series = chart.series[0]
        for idx, trade in enumerate(self.result.trades):
            pt = DataPoint(idx=idx)
            if trade.pnl_points >= 0:
                pt.graphicalProperties.solidFill = "00B050"
            else:
                pt.graphicalProperties.solidFill = "FF0000"
            series.data_points.append(pt)

        ws.add_chart(chart, "D2")
        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 5 – Monthly Breakdown
    # ------------------------------------------------------------------

    def _create_monthly_breakdown_sheet(self) -> None:
        ws = self.wb.create_sheet("Monthly Breakdown")

        headers = ["Month", "Trades", "Wins", "Losses", "Win Rate %", "Total P&L (pts)", "Avg P&L (pts)"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        monthly: dict[str, list[Trade]] = defaultdict(list)
        for trade in self.result.trades:
            if trade.entry_time:
                month_key = trade.entry_time.astimezone(IST).strftime("%Y-%m")
                monthly[month_key].append(trade)

        sorted_months = sorted(monthly.keys())
        for row_idx, month in enumerate(sorted_months, 2):
            trades = monthly[month]
            wins = sum(1 for t in trades if t.pnl_points > 0)
            losses = sum(1 for t in trades if t.pnl_points < 0)
            total_pnl = sum(t.pnl_points for t in trades)
            avg_pnl = total_pnl / len(trades) if trades else 0.0
            win_rate = (wins / len(trades)) * 100 if trades else 0.0

            ws.cell(row=row_idx, column=1, value=month)
            ws.cell(row=row_idx, column=2, value=len(trades))
            ws.cell(row=row_idx, column=3, value=wins)
            ws.cell(row=row_idx, column=4, value=losses)
            ws.cell(row=row_idx, column=5, value=round(win_rate, 2)).number_format = "0.00"

            pnl_cell = ws.cell(row=row_idx, column=6, value=round(total_pnl, 2))
            pnl_cell.number_format = "0.00"
            pnl_cell.fill = GREEN_FILL if total_pnl > 0 else RED_FILL if total_pnl < 0 else PatternFill()

            ws.cell(row=row_idx, column=7, value=round(avg_pnl, 2)).number_format = "0.00"

        if sorted_months:
            chart = BarChart()
            chart.title = "Monthly P&L"
            chart.x_axis.title = "Month"
            chart.y_axis.title = "P&L (points)"
            chart.width = 20
            chart.height = 12
            chart.style = 10

            data_ref = Reference(ws, min_col=6, min_row=1, max_row=len(sorted_months) + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_months) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            ws.add_chart(chart, "I2")

        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 6 – Daily P&L
    # ------------------------------------------------------------------

    def _create_daily_pnl_sheet(self) -> None:
        ws = self.wb.create_sheet("Daily P&L")

        headers = ["Date", "Trades", "Total P&L (pts)", "Cumulative P&L (pts)"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        daily: dict[str, list[Trade]] = defaultdict(list)
        for trade in self.result.trades:
            if trade.entry_time:
                day_key = trade.entry_time.astimezone(IST).strftime("%Y-%m-%d")
                daily[day_key].append(trade)

        sorted_days = sorted(daily.keys())
        cumulative = 0.0
        for row_idx, day in enumerate(sorted_days, 2):
            trades = daily[day]
            day_pnl = sum(t.pnl_points for t in trades)
            cumulative += day_pnl

            ws.cell(row=row_idx, column=1, value=day)
            ws.cell(row=row_idx, column=2, value=len(trades))

            pnl_cell = ws.cell(row=row_idx, column=3, value=round(day_pnl, 2))
            pnl_cell.number_format = "0.00"
            pnl_cell.fill = GREEN_FILL if day_pnl > 0 else RED_FILL if day_pnl < 0 else PatternFill()

            ws.cell(row=row_idx, column=4, value=round(cumulative, 2)).number_format = "0.00"

        if sorted_days:
            chart = LineChart()
            chart.title = "Daily Cumulative P&L"
            chart.x_axis.title = "Date"
            chart.y_axis.title = "Cumulative P&L (points)"
            chart.width = 20
            chart.height = 12
            chart.style = 10

            data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(sorted_days) + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_days) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.line.width = 20000

            ws.add_chart(chart, "F2")

        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 7 – Iron Condor Metrics (only for iron_condor strategy)
    # ------------------------------------------------------------------

    def _create_iron_condor_metrics_sheet(self) -> None:
        ws = self.wb.create_sheet("IC Metrics")
        r = self.result

        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = "Iron Condor Strategy Metrics"
        title_cell.font = TITLE_FONT

        row = 3

        def _metric(label: str, value, fmt: str | None = None) -> None:
            nonlocal row
            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            cell = ws.cell(row=row, column=2, value=value)
            if fmt:
                cell.number_format = fmt
            row += 1

        # Extract IC-specific metrics from trade metadata
        premiums: list[float] = []
        days_held: list[float] = []
        exit_reasons: dict[str, int] = {}
        spreads: list[int] = []
        lots_traded: list[int] = []

        for trade in r.trades:
            meta = getattr(trade, "metadata", {}) or {}
            if "max_profit" in meta:
                premiums.append(meta["max_profit"])
            if "spread_width" in meta:
                spreads.append(meta["spread_width"])
            reason = meta.get("reason", "unknown")
            exit_reasons[reason] = exit_reasons.get(reason, 0) + 1
            dh = meta.get("days_held", 0)
            if dh > 0:
                days_held.append(dh)
            qty = getattr(trade, "quantity", 0)
            if qty > 0:
                lots_traded.append(qty // 25)

        total = len(r.trades)

        _metric("Total Iron Condor Trades", total)
        row += 1

        # Premium metrics
        _metric("Avg Premium Collected (per unit)", 
                round(sum(premiums) / len(premiums), 2) if premiums else 0, "0.00")
        _metric("Max Premium Collected (per unit)",
                round(max(premiums), 2) if premiums else 0, "0.00")
        _metric("Min Premium Collected (per unit)",
                round(min(premiums), 2) if premiums else 0, "0.00")
        row += 1

        # Holding period
        _metric("Avg Days Held",
                round(sum(days_held) / len(days_held), 2) if days_held else 0, "0.00")
        _metric("Max Days Held",
                round(max(days_held), 2) if days_held else 0, "0.00")
        _metric("Min Days Held",
                round(min(days_held), 2) if days_held else 0, "0.00")
        row += 1

        # Lot sizing
        _metric("Avg Lots per Trade",
                round(sum(lots_traded) / len(lots_traded), 1) if lots_traded else 0, "0.0")
        _metric("Max Lots in Single Trade",
                max(lots_traded) if lots_traded else 0)
        row += 1

        # Exit reason breakdown
        _metric("EXIT REASON BREAKDOWN", "")
        for reason, count in sorted(exit_reasons.items(), key=lambda x: -x[1]):
            pct = (count / total * 100) if total > 0 else 0
            _metric(f"  {reason}", f"{count} ({pct:.1f}%)")

        row += 1

        # Brokerage impact
        _metric("Brokerage per Trade", f"₹{r.trades[0].brokerage:.0f}" if r.trades else "₹0")
        _metric("Total Brokerage", f"₹{r.total_brokerage:,.0f}")
        if r.total_pnl_rupees != 0:
            brokerage_pct = (r.total_brokerage / abs(r.total_pnl_rupees)) * 100
            _metric("Brokerage as % of Gross P&L", f"{brokerage_pct:.1f}%")

        _auto_width(ws)

    # ------------------------------------------------------------------
    # Sheet 8 – Expiry Cycle Breakdown (only for iron_condor strategy)
    # ------------------------------------------------------------------

    def _create_expiry_cycle_sheet(self) -> None:
        ws = self.wb.create_sheet("Expiry Cycles")

        headers = [
            "Expiry", "Trades", "Wins", "Losses", "Win Rate %",
            "Total P&L (pts)", "Total P&L (₹)", "Avg Days Held",
            "Profit Target", "Stop Loss", "EOD Exit", "Other",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Group trades by expiry from metadata
        by_expiry: dict[str, list[Trade]] = defaultdict(list)
        for trade in self.result.trades:
            meta = getattr(trade, "metadata", {}) or {}
            # Entry metadata has "expiry", exit metadata has "reason"
            expiry = meta.get("expiry", "unknown")
            if expiry == "unknown" and trade.entry_time:
                # Fallback: group by week of entry
                expiry = trade.entry_time.astimezone(IST).strftime("%Y-W%W")
            by_expiry[expiry].append(trade)

        sorted_expiries = sorted(by_expiry.keys())
        for row_idx, expiry in enumerate(sorted_expiries, 2):
            trades = by_expiry[expiry]
            wins = sum(1 for t in trades if t.pnl_points > 0)
            losses = sum(1 for t in trades if t.pnl_points < 0)
            total_pnl_pts = sum(t.pnl_points for t in trades)
            total_pnl_rs = sum(t.pnl_rupees for t in trades)
            win_rate = (wins / len(trades) * 100) if trades else 0

            # Days held from metadata
            dh_list = []
            exit_counts = {"profit_target": 0, "stop_loss": 0, "expiry_eod": 0, "other": 0}
            for t in trades:
                meta = getattr(t, "metadata", {}) or {}
                dh = meta.get("days_held", 0)
                if dh > 0:
                    dh_list.append(dh)
                reason = meta.get("reason", "other")
                if reason in exit_counts:
                    exit_counts[reason] += 1
                else:
                    exit_counts["other"] += 1

            avg_dh = sum(dh_list) / len(dh_list) if dh_list else 0

            ws.cell(row=row_idx, column=1, value=expiry)
            ws.cell(row=row_idx, column=2, value=len(trades))
            ws.cell(row=row_idx, column=3, value=wins)
            ws.cell(row=row_idx, column=4, value=losses)
            ws.cell(row=row_idx, column=5, value=round(win_rate, 1)).number_format = "0.0"

            pnl_cell = ws.cell(row=row_idx, column=6, value=round(total_pnl_pts, 2))
            pnl_cell.number_format = "0.00"
            pnl_cell.fill = GREEN_FILL if total_pnl_pts > 0 else RED_FILL if total_pnl_pts < 0 else PatternFill()

            rs_cell = ws.cell(row=row_idx, column=7, value=round(total_pnl_rs, 2))
            rs_cell.number_format = "#,##0.00"
            rs_cell.fill = GREEN_FILL if total_pnl_rs > 0 else RED_FILL if total_pnl_rs < 0 else PatternFill()

            ws.cell(row=row_idx, column=8, value=round(avg_dh, 1)).number_format = "0.0"
            ws.cell(row=row_idx, column=9, value=exit_counts["profit_target"])
            ws.cell(row=row_idx, column=10, value=exit_counts["stop_loss"])
            ws.cell(row=row_idx, column=11, value=exit_counts["expiry_eod"])
            ws.cell(row=row_idx, column=12, value=exit_counts["other"])

        # Chart: P&L by expiry cycle
        if sorted_expiries:
            chart = BarChart()
            chart.title = "P&L by Expiry Cycle"
            chart.x_axis.title = "Expiry"
            chart.y_axis.title = "P&L (₹)"
            chart.width = 20
            chart.height = 12
            chart.style = 10

            data_ref = Reference(ws, min_col=7, min_row=1, max_row=len(sorted_expiries) + 1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(sorted_expiries) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            ws.add_chart(chart, "N2")

        _auto_width(ws)
